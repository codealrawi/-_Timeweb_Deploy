"""
МедПлатформа API v3.1 — расширенная роль-модель + персональные рекомендации
- Роли: admin (управляет всем), doctor/patient (обычные пользователи)
- Реальные рекомендации на основе тегов постов и лайков
- 4 архитектуры для load-test (только админ)
"""
import os, time, hashlib, secrets, logging, asyncio
from io import BytesIO
from datetime import datetime
from typing import List, Optional, Dict, Set, Any
from collections import Counter, defaultdict

from fastapi import FastAPI, HTTPException, Depends
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field

import database as db
from services.moderation_service import ContentModerator
from services.recommendation_service import HybridRecommender, generate_demo_data
from services.architectures import get_architecture, list_architectures
from services.anomaly_service import (
    AccountAnomalyDetector,
    UserActivity,
    compute_features,
    generate_synthetic_activities,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("medplatforma")

app = FastAPI(title="МедПлатформа API", version="3.1.0")
app.add_middleware(
    CORSMiddleware, allow_origins=["*"],
    allow_credentials=True, allow_methods=["*"], allow_headers=["*"],
)

security = HTTPBearer(auto_error=False)

# ── In-memory хранилище ───────────────────────────────────────────────────────
_h = lambda p: hashlib.sha256(p.encode()).hexdigest()

MEM_USERS: Dict[str, dict] = {
    "admin": {"id":"admin","name":"Администратор","role":"admin","ph":_h("admin123"),"posts":0,"likes":0},
    "mod1":  {"id":"mod1","name":"Модератор Анна","role":"moderator","ph":_h("mod123"),"posts":0,"likes":0},
    "mod2":  {"id":"mod2","name":"Модератор Сергей","role":"moderator","ph":_h("mod123"),"posts":0,"likes":0},
    "doc1": {"id":"doc1","name":"Д-р Филиппов Д.Е.","role":"doctor","ph":_h("doctor123"),"posts":0,"likes":0},
    "doc2": {"id":"doc2","name":"Д-р Лебедев Д.А.","role":"doctor","ph":_h("doctor123"),"posts":0,"likes":0},
    "doc3": {"id":"doc3","name":"Д-р Сидоров В.А.","role":"doctor","ph":_h("doctor123"),"posts":0,"likes":0},
    "doc4": {"id":"doc4","name":"Д-р Орлов Р.Ф.","role":"doctor","ph":_h("doctor123"),"posts":0,"likes":0},
    "doc5": {"id":"doc5","name":"Д-р Воробьёв П.Б.","role":"doctor","ph":_h("doctor123"),"posts":0,"likes":0},
    "doc6": {"id":"doc6","name":"Д-р Давыдова П.В.","role":"doctor","ph":_h("doctor123"),"posts":0,"likes":0},
    "doc7": {"id":"doc7","name":"Д-р Голубева А.А.","role":"doctor","ph":_h("doctor123"),"posts":0,"likes":0},
    "doc8": {"id":"doc8","name":"Д-р Захарова О.М.","role":"doctor","ph":_h("doctor123"),"posts":0,"likes":0},
    "doc9": {"id":"doc9","name":"Д-р Волков Г.С.","role":"doctor","ph":_h("doctor123"),"posts":0,"likes":0},
    "doc10": {"id":"doc10","name":"Д-р Романов Н.А.","role":"doctor","ph":_h("doctor123"),"posts":0,"likes":0},
    "doc11": {"id":"doc11","name":"Д-р Петрова М.И.","role":"doctor","ph":_h("doctor123"),"posts":0,"likes":0},
    "doc12": {"id":"doc12","name":"Д-р Морозов Р.А.","role":"doctor","ph":_h("doctor123"),"posts":0,"likes":0},
    "pat1": {"id":"pat1","name":"Анна Николаева","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat2": {"id":"pat2","name":"Мария Лебедева","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat3": {"id":"pat3","name":"Светлана Петрова","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat4": {"id":"pat4","name":"Марина Степанова","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat5": {"id":"pat5","name":"Ольга Семёнова","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat6": {"id":"pat6","name":"Александр Егоров","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat7": {"id":"pat7","name":"Павел Новиков","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat8": {"id":"pat8","name":"Роман Козлов","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat9": {"id":"pat9","name":"Иван Волков","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat10": {"id":"pat10","name":"Виктория Егорова","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat11": {"id":"pat11","name":"Екатерина Сидорова","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat12": {"id":"pat12","name":"Дмитрий Алексеев","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat13": {"id":"pat13","name":"Полина Павлова","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat14": {"id":"pat14","name":"Артём Орлов","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat15": {"id":"pat15","name":"Николай Зайцев","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat16": {"id":"pat16","name":"Сергей Кузьмин","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat17": {"id":"pat17","name":"Дарья Новикова","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
    "pat18": {"id":"pat18","name":"Евгений Морозов","role":"patient","ph":_h("patient123"),"posts":0,"likes":0},
}
MEM_POSTS: List[dict] = [  # 90 постов
    {"id":"p54","author_id":"pat2","author":"Мария Лебедева","role":"patient",
     "title":"Ребёнок часто болеет ОРВИ",
     "body":"Сыну 4 года, болеет каждый месяц. Нужен ли иммунолог?",
     "tags":["педиатрия","иммунитет"],"likes_count":25,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p6","author_id":"doc2","author":"Д-р Лебедев Д.А.","role":"doctor",
     "title":"Целевой уровень HbA1c при диабете 2 типа",
     "body":"Для большинства пациентов цель HbA1c менее 7%. У пожилых — допустимо 7.5-8%.",
     "tags":["диабет","эндокринология"],"likes_count":35,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p47","author_id":"doc12","author":"Д-р Морозов Р.А.","role":"doctor",
     "title":"Когда можно сбивать температуру у ребёнка",
     "body":"До 38.5°C при хорошей переносимости — не сбиваем. Парацетамол или ибупрофен по весу ребёнка.",
     "tags":["педиатрия","препараты"],"likes_count":80,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p56","author_id":"pat2","author":"Мария Лебедева","role":"patient",
     "title":"Как правильно измерять давление дома?",
     "body":"Купил тонометр. Утром одно значение, вечером другое.",
     "tags":["кардиология","гипертония"],"likes_count":20,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p19","author_id":"doc5","author":"Д-р Воробьёв П.Б.","role":"doctor",
     "title":"Реабилитация после инфаркта миокарда",
     "body":"После инфаркта важна кардиореабилитация. Первые 6 недель — ограниченная физическая активность.",
     "tags":["кардиология","реабилитация"],"likes_count":52,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p44","author_id":"doc11","author":"Д-р Петрова М.И.","role":"doctor",
     "title":"Когда можно сбивать температуру у ребёнка",
     "body":"До 38.5°C при хорошей переносимости — не сбиваем. Парацетамол или ибупрофен по весу ребёнка.",
     "tags":["педиатрия","препараты"],"likes_count":12,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p83","author_id":"pat14","author":"Артём Орлов","role":"patient",
     "title":"Можно ли совмещать ибупрофен и парацетамол?",
     "body":"При высокой температуре врач сказал чередовать. Через сколько часов?",
     "tags":["терапия","препараты"],"likes_count":26,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p10","author_id":"doc3","author":"Д-р Сидоров В.А.","role":"doctor",
     "title":"Аритмия: когда стоит беспокоиться",
     "body":"Экстрасистолы у здоровых людей бывают, но при частоте более 6/мин — нужна консультация кардиолога.",
     "tags":["кардиология","аритмия"],"likes_count":36,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p7","author_id":"doc3","author":"Д-р Сидоров В.А.","role":"doctor",
     "title":"Гастрит: диета и лечение",
     "body":"При обострении — стол №1 по Певзнеру. При Helicobacter pylori — эрадикационная терапия 14 дней.",
     "tags":["гастроэнтерология","питание"],"likes_count":85,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p49","author_id":"doc12","author":"Д-р Морозов Р.А.","role":"doctor",
     "title":"Когда можно сбивать температуру у ребёнка",
     "body":"До 38.5°C при хорошей переносимости — не сбиваем. Парацетамол или ибупрофен по весу ребёнка.",
     "tags":["педиатрия","препараты"],"likes_count":71,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p8","author_id":"doc3","author":"Д-р Сидоров В.А.","role":"doctor",
     "title":"Артериальная гипертензия: когда начинать лечение",
     "body":"Диагноз артериальной гипертензии ставится при стойком повышении АД выше 140/90 мм рт. ст. Немедикаментозное лечение показано всем пациентам.",
     "tags":["кардиология","гипертония"],"likes_count":79,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p39","author_id":"doc10","author":"Д-р Романов Н.А.","role":"doctor",
     "title":"Мигрень: триггеры и профилактика",
     "body":"Сон, питание, стресс — основные триггеры. При частых приступах назначается профилактическая терапия.",
     "tags":["неврология","мигрень"],"likes_count":67,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p25","author_id":"doc6","author":"Д-р Давыдова П.В.","role":"doctor",
     "title":"Антибиотики при пневмонии",
     "body":"Амоксициллин/клавуланат как первая линия. Курс 7-10 дней. При атипичной пневмонии — макролиды.",
     "tags":["пульмонология","антибиотики"],"likes_count":88,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p24","author_id":"doc6","author":"Д-р Давыдова П.В.","role":"doctor",
     "title":"Целевой уровень HbA1c при диабете 2 типа",
     "body":"Для большинства пациентов цель HbA1c менее 7%. У пожилых — допустимо 7.5-8%.",
     "tags":["диабет","эндокринология"],"likes_count":78,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p90","author_id":"pat18","author":"Евгений Морозов","role":"patient",
     "title":"Боли в спине при сидячей работе",
     "body":"Работаю за компьютером 8 часов. К концу дня ноет поясница. Какие упражнения помогут?",
     "tags":["ортопедия","остеохондроз","лфк"],"likes_count":30,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p20","author_id":"doc5","author":"Д-р Воробьёв П.Б.","role":"doctor",
     "title":"Остеохондроз шейного отдела",
     "body":"ЛФК — основа лечения. НПВС короткими курсами при обострении. МРТ при корешковом синдроме.",
     "tags":["ортопедия","остеохондроз","лфк"],"likes_count":41,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p88","author_id":"pat17","author":"Дарья Новикова","role":"patient",
     "title":"Как правильно измерять давление дома?",
     "body":"Купил тонометр. Утром одно значение, вечером другое.",
     "tags":["кардиология","гипертония"],"likes_count":13,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p42","author_id":"doc11","author":"Д-р Петрова М.И.","role":"doctor",
     "title":"Профилактика ОРВИ у детей",
     "body":"Закаливание, проветривание, промывание носа физраствором. Вакцинация от гриппа ежегодно.",
     "tags":["педиатрия","профилактика","вакцинация"],"likes_count":49,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p23","author_id":"doc6","author":"Д-р Давыдова П.В.","role":"doctor",
     "title":"Бронхиальная астма: базисная терапия",
     "body":"Ингаляционные кортикостероиды — основа лечения. Сальбутамол — только для купирования приступов.",
     "tags":["пульмонология","астма"],"likes_count":56,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p30","author_id":"doc7","author":"Д-р Голубева А.А.","role":"doctor",
     "title":"Целевой уровень HbA1c при диабете 2 типа",
     "body":"Для большинства пациентов цель HbA1c менее 7%. У пожилых — допустимо 7.5-8%.",
     "tags":["диабет","эндокринология"],"likes_count":37,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p21","author_id":"doc5","author":"Д-р Воробьёв П.Б.","role":"doctor",
     "title":"Депрессия: когда нужна помощь",
     "body":"Снижение настроения более 2 недель — повод обратиться к психиатру. СИОЗС — первая линия.",
     "tags":["психиатрия","депрессия"],"likes_count":13,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p65","author_id":"pat8","author":"Роман Козлов","role":"patient",
     "title":"Сильная изжога после еды",
     "body":"Изжога каждый день после обеда и ужина. Нужна ли гастроскопия?",
     "tags":["гастроэнтерология","терапия"],"likes_count":18,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p62","author_id":"pat6","author":"Александр Егоров","role":"patient",
     "title":"Как снизить сахар без таблеток?",
     "body":"Сахар натощак 6.5. Врач пока не назначает препараты. Какая диета?",
     "tags":["диабет","питание"],"likes_count":23,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p51","author_id":"doc12","author":"Д-р Морозов Р.А.","role":"doctor",
     "title":"Гипотиреоз: симптомы и лечение",
     "body":"Усталость, зябкость, прибавка веса — повод сдать ТТГ. При повышенном ТТГ назначается L-тироксин.",
     "tags":["эндокринология"],"likes_count":26,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p74","author_id":"pat11","author":"Екатерина Сидорова","role":"patient",
     "title":"Сильная изжога после еды",
     "body":"Изжога каждый день после обеда и ужина. Нужна ли гастроскопия?",
     "tags":["гастроэнтерология","терапия"],"likes_count":2,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p76","author_id":"pat12","author":"Дмитрий Алексеев","role":"patient",
     "title":"Ребёнок часто болеет ОРВИ",
     "body":"Сыну 4 года, болеет каждый месяц. Нужен ли иммунолог?",
     "tags":["педиатрия","иммунитет"],"likes_count":18,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p12","author_id":"doc4","author":"Д-р Орлов Р.Ф.","role":"doctor",
     "title":"Депрессия: когда нужна помощь",
     "body":"Снижение настроения более 2 недель — повод обратиться к психиатру. СИОЗС — первая линия.",
     "tags":["психиатрия","депрессия"],"likes_count":8,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p53","author_id":"pat1","author":"Анна Николаева","role":"patient",
     "title":"Красные пятна на коже",
     "body":"Появились пятна на руках и ногах, чешутся. К кому идти?",
     "tags":["дерматология","аллергия"],"likes_count":19,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p81","author_id":"pat14","author":"Артём Орлов","role":"patient",
     "title":"Тревожность и плохой сон",
     "body":"Не могу уснуть, мысли крутятся. Нужен ли мне психиатр?",
     "tags":["психиатрия","сон","стресс"],"likes_count":21,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p79","author_id":"pat13","author":"Полина Павлова","role":"patient",
     "title":"Онемение пальцев рук по утрам",
     "body":"Просыпаюсь с онемевшими пальцами. Через 10 минут проходит.",
     "tags":["неврология","диагностика"],"likes_count":22,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p72","author_id":"pat10","author":"Виктория Егорова","role":"patient",
     "title":"Можно ли совмещать ибупрофен и парацетамол?",
     "body":"При высокой температуре врач сказал чередовать. Через сколько часов?",
     "tags":["терапия","препараты"],"likes_count":5,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p13","author_id":"doc4","author":"Д-р Орлов Р.Ф.","role":"doctor",
     "title":"Депрессия: когда нужна помощь",
     "body":"Снижение настроения более 2 недель — повод обратиться к психиатру. СИОЗС — первая линия.",
     "tags":["психиатрия","депрессия"],"likes_count":28,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p27","author_id":"doc6","author":"Д-р Давыдова П.В.","role":"doctor",
     "title":"Остеохондроз шейного отдела",
     "body":"ЛФК — основа лечения. НПВС короткими курсами при обострении. МРТ при корешковом синдроме.",
     "tags":["ортопедия","остеохондроз","лфк"],"likes_count":32,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p60","author_id":"pat4","author":"Марина Степанова","role":"patient",
     "title":"Высокое давление по утрам — что делать?",
     "body":"Каждое утро АД около 150/95. Принимаю эналаприл 10 мг. Эффекта почти нет.",
     "tags":["кардиология","гипертония"],"likes_count":29,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p45","author_id":"doc11","author":"Д-р Петрова М.И.","role":"doctor",
     "title":"Депрессия: когда нужна помощь",
     "body":"Снижение настроения более 2 недель — повод обратиться к психиатру. СИОЗС — первая линия.",
     "tags":["психиатрия","депрессия"],"likes_count":48,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p31","author_id":"doc8","author":"Д-р Захарова О.М.","role":"doctor",
     "title":"Аритмия: когда стоит беспокоиться",
     "body":"Экстрасистолы у здоровых людей бывают, но при частоте более 6/мин — нужна консультация кардиолога.",
     "tags":["кардиология","аритмия"],"likes_count":43,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p78","author_id":"pat13","author":"Полина Павлова","role":"patient",
     "title":"Как правильно измерять давление дома?",
     "body":"Купил тонометр. Утром одно значение, вечером другое.",
     "tags":["кардиология","гипертония"],"likes_count":5,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p37","author_id":"doc9","author":"Д-р Волков Г.С.","role":"doctor",
     "title":"Целевой уровень HbA1c при диабете 2 типа",
     "body":"Для большинства пациентов цель HbA1c менее 7%. У пожилых — допустимо 7.5-8%.",
     "tags":["диабет","эндокринология"],"likes_count":85,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p64","author_id":"pat7","author":"Павел Новиков","role":"patient",
     "title":"Какой анализ на щитовидку сдать?",
     "body":"Чувствую усталость, набираю вес. С чего начать обследование?",
     "tags":["эндокринология","анализы"],"likes_count":14,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p35","author_id":"doc9","author":"Д-р Волков Г.С.","role":"doctor",
     "title":"Когда можно сбивать температуру у ребёнка",
     "body":"До 38.5°C при хорошей переносимости — не сбиваем. Парацетамол или ибупрофен по весу ребёнка.",
     "tags":["педиатрия","препараты"],"likes_count":42,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p14","author_id":"doc4","author":"Д-р Орлов Р.Ф.","role":"doctor",
     "title":"Витамин D: кому и сколько",
     "body":"Профилактическая доза 1000-2000 МЕ/сут. При дефиците — лечебные дозы. Контроль через 3 месяца.",
     "tags":["эндокринология","витамины"],"likes_count":62,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p66","author_id":"pat8","author":"Роман Козлов","role":"patient",
     "title":"Ребёнок часто болеет ОРВИ",
     "body":"Сыну 4 года, болеет каждый месяц. Нужен ли иммунолог?",
     "tags":["педиатрия","иммунитет"],"likes_count":19,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p43","author_id":"doc11","author":"Д-р Петрова М.И.","role":"doctor",
     "title":"Варикозное расширение вен",
     "body":"Компрессионный трикотаж — основа лечения. При прогрессировании — консультация флеболога.",
     "tags":["хирургия","терапия"],"likes_count":15,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p18","author_id":"doc5","author":"Д-р Воробьёв П.Б.","role":"doctor",
     "title":"Аритмия: когда стоит беспокоиться",
     "body":"Экстрасистолы у здоровых людей бывают, но при частоте более 6/мин — нужна консультация кардиолога.",
     "tags":["кардиология","аритмия"],"likes_count":20,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p85","author_id":"pat16","author":"Сергей Кузьмин","role":"patient",
     "title":"Как правильно измерять давление дома?",
     "body":"Купил тонометр. Утром одно значение, вечером другое.",
     "tags":["кардиология","гипертония"],"likes_count":2,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p33","author_id":"doc8","author":"Д-р Захарова О.М.","role":"doctor",
     "title":"Варикозное расширение вен",
     "body":"Компрессионный трикотаж — основа лечения. При прогрессировании — консультация флеболога.",
     "tags":["хирургия","терапия"],"likes_count":54,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p50","author_id":"doc12","author":"Д-р Морозов Р.А.","role":"doctor",
     "title":"Аритмия: когда стоит беспокоиться",
     "body":"Экстрасистолы у здоровых людей бывают, но при частоте более 6/мин — нужна консультация кардиолога.",
     "tags":["кардиология","аритмия"],"likes_count":90,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p87","author_id":"pat16","author":"Сергей Кузьмин","role":"patient",
     "title":"Сильная изжога после еды",
     "body":"Изжога каждый день после обеда и ужина. Нужна ли гастроскопия?",
     "tags":["гастроэнтерология","терапия"],"likes_count":2,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p3","author_id":"doc1","author":"Д-р Филиппов Д.Е.","role":"doctor",
     "title":"Мигрень: триггеры и профилактика",
     "body":"Сон, питание, стресс — основные триггеры. При частых приступах назначается профилактическая терапия.",
     "tags":["неврология","мигрень"],"likes_count":21,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p77","author_id":"pat12","author":"Дмитрий Алексеев","role":"patient",
     "title":"Как снизить сахар без таблеток?",
     "body":"Сахар натощак 6.5. Врач пока не назначает препараты. Какая диета?",
     "tags":["диабет","питание"],"likes_count":7,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p75","author_id":"pat12","author":"Дмитрий Алексеев","role":"patient",
     "title":"Как правильно измерять давление дома?",
     "body":"Купил тонометр. Утром одно значение, вечером другое.",
     "tags":["кардиология","гипертония"],"likes_count":25,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p4","author_id":"doc2","author":"Д-р Лебедев Д.А.","role":"doctor",
     "title":"Остеохондроз шейного отдела",
     "body":"ЛФК — основа лечения. НПВС короткими курсами при обострении. МРТ при корешковом синдроме.",
     "tags":["ортопедия","остеохондроз","лфк"],"likes_count":62,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p41","author_id":"doc10","author":"Д-р Романов Н.А.","role":"doctor",
     "title":"Бронхиальная астма: базисная терапия",
     "body":"Ингаляционные кортикостероиды — основа лечения. Сальбутамол — только для купирования приступов.",
     "tags":["пульмонология","астма"],"likes_count":89,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p46","author_id":"doc12","author":"Д-р Морозов Р.А.","role":"doctor",
     "title":"Антибиотики при пневмонии",
     "body":"Амоксициллин/клавуланат как первая линия. Курс 7-10 дней. При атипичной пневмонии — макролиды.",
     "tags":["пульмонология","антибиотики"],"likes_count":16,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p5","author_id":"doc2","author":"Д-р Лебедев Д.А.","role":"doctor",
     "title":"Артериальная гипертензия: когда начинать лечение",
     "body":"Диагноз артериальной гипертензии ставится при стойком повышении АД выше 140/90 мм рт. ст. Немедикаментозное лечение показано всем пациентам.",
     "tags":["кардиология","гипертония"],"likes_count":11,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p1","author_id":"doc1","author":"Д-р Филиппов Д.Е.","role":"doctor",
     "title":"Артериальная гипертензия: когда начинать лечение",
     "body":"Диагноз артериальной гипертензии ставится при стойком повышении АД выше 140/90 мм рт. ст. Немедикаментозное лечение показано всем пациентам.",
     "tags":["кардиология","гипертония"],"likes_count":43,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p59","author_id":"pat4","author":"Марина Степанова","role":"patient",
     "title":"Как подготовиться к УЗИ брюшной полости?",
     "body":"Назначено УЗИ на утро. Что можно есть накануне? Нужно ли голодать?",
     "tags":["диагностика","узи"],"likes_count":26,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p82","author_id":"pat14","author":"Артём Орлов","role":"patient",
     "title":"Боли в спине при сидячей работе",
     "body":"Работаю за компьютером 8 часов. К концу дня ноет поясница. Какие упражнения помогут?",
     "tags":["ортопедия","остеохондроз","лфк"],"likes_count":6,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p15","author_id":"doc4","author":"Д-р Орлов Р.Ф.","role":"doctor",
     "title":"Реабилитация после инфаркта миокарда",
     "body":"После инфаркта важна кардиореабилитация. Первые 6 недель — ограниченная физическая активность.",
     "tags":["кардиология","реабилитация"],"likes_count":43,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p36","author_id":"doc9","author":"Д-р Волков Г.С.","role":"doctor",
     "title":"Витамин D: кому и сколько",
     "body":"Профилактическая доза 1000-2000 МЕ/сут. При дефиците — лечебные дозы. Контроль через 3 месяца.",
     "tags":["эндокринология","витамины"],"likes_count":90,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p38","author_id":"doc10","author":"Д-р Романов Н.А.","role":"doctor",
     "title":"Гастрит: диета и лечение",
     "body":"При обострении — стол №1 по Певзнеру. При Helicobacter pylori — эрадикационная терапия 14 дней.",
     "tags":["гастроэнтерология","питание"],"likes_count":39,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p2","author_id":"doc1","author":"Д-р Филиппов Д.Е.","role":"doctor",
     "title":"Когда можно сбивать температуру у ребёнка",
     "body":"До 38.5°C при хорошей переносимости — не сбиваем. Парацетамол или ибупрофен по весу ребёнка.",
     "tags":["педиатрия","препараты"],"likes_count":36,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p80","author_id":"pat13","author":"Полина Павлова","role":"patient",
     "title":"Ребёнок часто болеет ОРВИ",
     "body":"Сыну 4 года, болеет каждый месяц. Нужен ли иммунолог?",
     "tags":["педиатрия","иммунитет"],"likes_count":28,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p86","author_id":"pat16","author":"Сергей Кузьмин","role":"patient",
     "title":"Болит колено при ходьбе по лестнице",
     "body":"Хруст при сгибании. Нужен рентген или МРТ?",
     "tags":["ортопедия","диагностика"],"likes_count":12,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p29","author_id":"doc7","author":"Д-р Голубева А.А.","role":"doctor",
     "title":"Депрессия: когда нужна помощь",
     "body":"Снижение настроения более 2 недель — повод обратиться к психиатру. СИОЗС — первая линия.",
     "tags":["психиатрия","депрессия"],"likes_count":45,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p32","author_id":"doc8","author":"Д-р Захарова О.М.","role":"doctor",
     "title":"Гипотиреоз: симптомы и лечение",
     "body":"Усталость, зябкость, прибавка веса — повод сдать ТТГ. При повышенном ТТГ назначается L-тироксин.",
     "tags":["эндокринология"],"likes_count":89,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p16","author_id":"doc4","author":"Д-р Орлов Р.Ф.","role":"doctor",
     "title":"Мигрень: триггеры и профилактика",
     "body":"Сон, питание, стресс — основные триггеры. При частых приступах назначается профилактическая терапия.",
     "tags":["неврология","мигрень"],"likes_count":35,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p58","author_id":"pat3","author":"Светлана Петрова","role":"patient",
     "title":"Частые головные боли по вечерам",
     "body":"Две недели к вечеру болит голова в области висков. Какие обследования пройти?",
     "tags":["неврология","мигрень"],"likes_count":18,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p67","author_id":"pat8","author":"Роман Козлов","role":"patient",
     "title":"Онемение пальцев рук по утрам",
     "body":"Просыпаюсь с онемевшими пальцами. Через 10 минут проходит.",
     "tags":["неврология","диагностика"],"likes_count":2,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p57","author_id":"pat3","author":"Светлана Петрова","role":"patient",
     "title":"Можно ли совмещать ибупрофен и парацетамол?",
     "body":"При высокой температуре врач сказал чередовать. Через сколько часов?",
     "tags":["терапия","препараты"],"likes_count":9,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p48","author_id":"doc12","author":"Д-р Морозов Р.А.","role":"doctor",
     "title":"Бронхиальная астма: базисная терапия",
     "body":"Ингаляционные кортикостероиды — основа лечения. Сальбутамол — только для купирования приступов.",
     "tags":["пульмонология","астма"],"likes_count":48,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p52","author_id":"pat1","author":"Анна Николаева","role":"patient",
     "title":"Частые головные боли по вечерам",
     "body":"Две недели к вечеру болит голова в области висков. Какие обследования пройти?",
     "tags":["неврология","мигрень"],"likes_count":9,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p40","author_id":"doc10","author":"Д-р Романов Н.А.","role":"doctor",
     "title":"Аритмия: когда стоит беспокоиться",
     "body":"Экстрасистолы у здоровых людей бывают, но при частоте более 6/мин — нужна консультация кардиолога.",
     "tags":["кардиология","аритмия"],"likes_count":42,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p26","author_id":"doc6","author":"Д-р Давыдова П.В.","role":"doctor",
     "title":"Остеохондроз шейного отдела",
     "body":"ЛФК — основа лечения. НПВС короткими курсами при обострении. МРТ при корешковом синдроме.",
     "tags":["ортопедия","остеохондроз","лфк"],"likes_count":54,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p70","author_id":"pat9","author":"Иван Волков","role":"patient",
     "title":"Тревожность и плохой сон",
     "body":"Не могу уснуть, мысли крутятся. Нужен ли мне психиатр?",
     "tags":["психиатрия","сон","стресс"],"likes_count":26,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p28","author_id":"doc7","author":"Д-р Голубева А.А.","role":"doctor",
     "title":"Артериальная гипертензия: когда начинать лечение",
     "body":"Диагноз артериальной гипертензии ставится при стойком повышении АД выше 140/90 мм рт. ст. Немедикаментозное лечение показано всем пациентам.",
     "tags":["кардиология","гипертония"],"likes_count":37,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p55","author_id":"pat2","author":"Мария Лебедева","role":"patient",
     "title":"Болит колено при ходьбе по лестнице",
     "body":"Хруст при сгибании. Нужен рентген или МРТ?",
     "tags":["ортопедия","диагностика"],"likes_count":15,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p68","author_id":"pat9","author":"Иван Волков","role":"patient",
     "title":"Красные пятна на коже",
     "body":"Появились пятна на руках и ногах, чешутся. К кому идти?",
     "tags":["дерматология","аллергия"],"likes_count":5,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p34","author_id":"doc9","author":"Д-р Волков Г.С.","role":"doctor",
     "title":"Реабилитация после инфаркта миокарда",
     "body":"После инфаркта важна кардиореабилитация. Первые 6 недель — ограниченная физическая активность.",
     "tags":["кардиология","реабилитация"],"likes_count":53,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p22","author_id":"doc6","author":"Д-р Давыдова П.В.","role":"doctor",
     "title":"Гастрит: диета и лечение",
     "body":"При обострении — стол №1 по Певзнеру. При Helicobacter pylori — эрадикационная терапия 14 дней.",
     "tags":["гастроэнтерология","питание"],"likes_count":23,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p71","author_id":"pat10","author":"Виктория Егорова","role":"patient",
     "title":"Как снизить сахар без таблеток?",
     "body":"Сахар натощак 6.5. Врач пока не назначает препараты. Какая диета?",
     "tags":["диабет","питание"],"likes_count":22,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p61","author_id":"pat5","author":"Ольга Семёнова","role":"patient",
     "title":"Частые головные боли по вечерам",
     "body":"Две недели к вечеру болит голова в области висков. Какие обследования пройти?",
     "tags":["неврология","мигрень"],"likes_count":22,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p84","author_id":"pat15","author":"Николай Зайцев","role":"patient",
     "title":"Тревожность и плохой сон",
     "body":"Не могу уснуть, мысли крутятся. Нужен ли мне психиатр?",
     "tags":["психиатрия","сон","стресс"],"likes_count":26,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p17","author_id":"doc5","author":"Д-р Воробьёв П.Б.","role":"doctor",
     "title":"Целевой уровень HbA1c при диабете 2 типа",
     "body":"Для большинства пациентов цель HbA1c менее 7%. У пожилых — допустимо 7.5-8%.",
     "tags":["диабет","эндокринология"],"likes_count":19,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p69","author_id":"pat9","author":"Иван Волков","role":"patient",
     "title":"Повышенный холестерин — нужны ли статины?",
     "body":"Общий холестерин 6.8, ЛПНП 4.2. Есть ли альтернативы?",
     "tags":["кардиология","анализы","питание"],"likes_count":30,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p9","author_id":"doc3","author":"Д-р Сидоров В.А.","role":"doctor",
     "title":"Когда можно сбивать температуру у ребёнка",
     "body":"До 38.5°C при хорошей переносимости — не сбиваем. Парацетамол или ибупрофен по весу ребёнка.",
     "tags":["педиатрия","препараты"],"likes_count":77,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p63","author_id":"pat7","author":"Павел Новиков","role":"patient",
     "title":"Болит колено при ходьбе по лестнице",
     "body":"Хруст при сгибании. Нужен рентген или МРТ?",
     "tags":["ортопедия","диагностика"],"likes_count":4,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p89","author_id":"pat18","author":"Евгений Морозов","role":"patient",
     "title":"Боли в спине при сидячей работе",
     "body":"Работаю за компьютером 8 часов. К концу дня ноет поясница. Какие упражнения помогут?",
     "tags":["ортопедия","остеохондроз","лфк"],"likes_count":3,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p11","author_id":"doc4","author":"Д-р Орлов Р.Ф.","role":"doctor",
     "title":"Остеохондроз шейного отдела",
     "body":"ЛФК — основа лечения. НПВС короткими курсами при обострении. МРТ при корешковом синдроме.",
     "tags":["ортопедия","остеохондроз","лфк"],"likes_count":43,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
    {"id":"p73","author_id":"pat11","author":"Екатерина Сидорова","role":"patient",
     "title":"Какой анализ на щитовидку сдать?",
     "body":"Чувствую усталость, набираю вес. С чего начать обследование?",
     "tags":["эндокринология","анализы"],"likes_count":7,"status":"approved",
     "created_at":datetime.utcnow().isoformat()},
]
MEM_LIKES: Dict[str, Set[str]] = defaultdict(set)
_TOKENS: dict = {}
_NEXT_USER_ID = [300]
_NEXT_POST_ID = [191]  # счётчик id для новых постов

# ── Жизненный цикл ────────────────────────────────────────────────────────────
@app.on_event("startup")
async def startup():
    app.state.mod = ContentModerator()
    app.state.mod.train()
    items, ints = generate_demo_data()
    app.state.rec = HybridRecommender(alpha=0.4)
    app.state.rec.fit(items, ints)
    logger.info("[INIT] ML-сервисы готовы")
    app.state.use_db = await db.init_pool()
    logger.info("[INIT] БД: %s",
                "postgresql" if app.state.use_db else "in-memory (БД недоступна)")

@app.on_event("shutdown")
async def shutdown():
    await db.close_pool()

# ── Статика ───────────────────────────────────────────────────────────────────
if os.path.isdir("static"):
    app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/app", include_in_schema=False)
async def app_index():
    idx = "static/index.html"
    if os.path.exists(idx):
        return FileResponse(idx)
    from fastapi.responses import HTMLResponse
    return HTMLResponse("<h2>🏥 МедПлатформа — фронтенд не найден</h2>", 404)

# ── Модели ────────────────────────────────────────────────────────────────────
class LoginReq(BaseModel):
    username: str
    password: str

class PostReq(BaseModel):
    title: str = Field(..., min_length=5, max_length=500)
    body:  str = Field(..., min_length=10)
    tags:  List[str] = []

class PostUpdateReq(BaseModel):
    title: Optional[str] = Field(None, min_length=5, max_length=500)
    body:  Optional[str] = Field(None, min_length=10)
    tags:  Optional[List[str]] = None
    status: Optional[str] = None

class ModReq(BaseModel):
    text: str

class UserCreateReq(BaseModel):
    username: str = Field(..., min_length=2, max_length=30)
    password: str = Field(..., min_length=4, max_length=100)
    name: str = Field(..., min_length=2, max_length=100)
    role: str = Field(..., pattern="^(doctor|patient|admin|moderator)$")

class UserUpdateReq(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = Field(None, pattern="^(doctor|patient|admin|moderator)$")
    password: Optional[str] = None

class AnomalyDetectReq(BaseModel):
    synthetic: bool = False
    n_synthetic_legit: int = Field(45, ge=0, le=200)
    n_synthetic_anomalies: int = Field(5, ge=0, le=50)
    update_db: bool = True

class PostReviewReq(BaseModel):
    """Запрос модератора на обработку подозрительного поста."""
    action: str = Field(..., pattern="^(approve|reject|recheck)$")
    title: Optional[str] = None     # редактированный заголовок
    body: Optional[str] = None      # редактированное тело
    comment: Optional[str] = None   # комментарий модератора

class XlsxSheet(BaseModel):
    name: str
    headers: List[str] = []
    rows: List[List[Any]] = []

class XlsxExportReq(BaseModel):
    filename: str = "export"
    title: Optional[str] = None
    sheets: List[XlsxSheet]

# ── Авторизация и роли ────────────────────────────────────────────────────────
def get_user(creds: HTTPAuthorizationCredentials = Depends(security)):
    if not creds or creds.credentials not in _TOKENS:
        raise HTTPException(401, "Не авторизован")
    return _TOKENS[creds.credentials]

def get_user_optional(creds: HTTPAuthorizationCredentials = Depends(security)):
    if not creds or creds.credentials not in _TOKENS:
        return None
    return _TOKENS[creds.credentials]

def require_admin(user=Depends(get_user)):
    if user.get("role") != "admin":
        raise HTTPException(403, "Доступ только для администратора")
    return user

def require_moderator(user=Depends(get_user)):
    """Модератор ИЛИ администратор — доступ к модерации и проверке постов."""
    if user.get("role") not in ("moderator", "admin"):
        raise HTTPException(403, "Доступ для модератора или администратора")
    return user

# ── Базовые эндпоинты ─────────────────────────────────────────────────────────
@app.get("/")
async def root():
    return {
        "service":   "МедПлатформа API",
        "version":   "3.1.0",
        "status":    "running",
        "database":  "postgresql" if app.state.use_db else "in-memory",
        "docs":      "/docs",
        "app":       "/app",
        "timestamp": datetime.utcnow().isoformat(),
    }

@app.get("/health")
async def health():
    db_status = await db.db_ok() if app.state.use_db else False
    return {
        "status":        "ok",
        "database":      db_status,
        "database_mode": "postgresql" if app.state.use_db else "in-memory",
        "moderator":     True,
        "recommender":   True,
        "timestamp":     datetime.utcnow().isoformat(),
    }

@app.post("/auth/login")
async def login(req: LoginReq):
    ph = hashlib.sha256(req.password.encode()).hexdigest()
    if app.state.use_db and db.pool:
        row = await db.pool.fetchrow(
            "SELECT id, name, role FROM users WHERE id=$1 AND password_hash=$2",
            req.username, ph,
        )
        if not row:
            raise HTTPException(401, "Неверный логин или пароль")
        user = dict(row)
    else:
        u = MEM_USERS.get(req.username)
        if not u or u["ph"] != ph:
            raise HTTPException(401, "Неверный логин или пароль")
        user = {"id": u["id"], "name": u["name"], "role": u["role"]}

    tok = secrets.token_hex(32)
    _TOKENS[tok] = user
    return {"token": tok, "user": user}

# ── Посты ─────────────────────────────────────────────────────────────────────
@app.get("/posts")
async def get_posts(limit: int = 20, offset: int = 0,
                    user=Depends(get_user_optional)):
    user_id = user["id"] if user else None

    if app.state.use_db and db.pool:
        rows = await db.pool.fetch("""
            SELECT p.id, p.author_id, p.title, p.body, p.status, p.likes_count,
                   u.name AS author, u.role,
                   array_agg(t.name) FILTER (WHERE t.name IS NOT NULL) AS tags,
                   EXISTS(SELECT 1 FROM likes l WHERE l.post_id=p.id AND l.user_id=$3) AS liked_by_me
            FROM posts p
            JOIN users u ON u.id = p.author_id
            LEFT JOIN post_tags pt ON pt.post_id = p.id
            LEFT JOIN tags t ON t.id = pt.tag_id
            WHERE p.status = 'approved'
            GROUP BY p.id, u.id
            ORDER BY p.created_at DESC
            LIMIT $1 OFFSET $2
        """, limit, offset, user_id or "")
        total = await db.pool.fetchval(
            "SELECT COUNT(*) FROM posts WHERE status='approved'")
        return {"posts": [dict(r) for r in rows], "total": total}

    # in-memory: с фильтром approved + флаг liked_by_me
    filtered = [p for p in MEM_POSTS if p.get("status") == "approved"]
    sl = filtered[offset:offset+limit]
    out = []
    for p in sl:
        d = dict(p)
        d["liked_by_me"] = bool(user_id) and p["id"] in MEM_LIKES.get(user_id, set())
        out.append(d)
    return {"posts": out, "total": len(filtered)}

@app.post("/posts", status_code=201)
async def create_post(req: PostReq, user=Depends(get_user)):
    full_text = f"{req.title}\n{req.body}"
    mod = app.state.mod.moderate(full_text)

    # ── Publish-flow по вердикту модерации ──
    # BLOCKED → отказ, пост НЕ создаётся
    if mod.label == "blocked":
        raise HTTPException(422, detail={
            "message": "Текст заблокирован — публикация невозможна",
            "moderation": {
                "label": mod.label, "confidence": mod.confidence,
                "level": mod.level, "reasons": mod.reasons,
            },
        })

    # SUSPICIOUS → пост создаётся со статусом 'suspicious', не виден в ленте,
    # попадает в очередь модератора
    # APPROVED → пост создаётся со статусом 'approved', сразу в ленту
    status = "suspicious" if mod.label == "suspicious" else "approved"

    pid = f"p{_NEXT_POST_ID[0]}"
    _NEXT_POST_ID[0] += 1

    if app.state.use_db and db.pool:
        async with db.pool.acquire() as conn:
            async with conn.transaction():
                await conn.execute(
                    "INSERT INTO posts(id,author_id,title,body,status,mod_level,mod_conf)"
                    " VALUES($1,$2,$3,$4,$5,$6,$7)",
                    pid, user["id"], req.title, req.body,
                    status, mod.level, float(mod.confidence),
                )
                for tag in req.tags[:10]:
                    tag = tag.strip().lower()
                    if tag:
                        tid = await conn.fetchval(
                            "INSERT INTO tags(name) VALUES($1)"
                            " ON CONFLICT(name) DO UPDATE SET name=EXCLUDED.name RETURNING id",
                            tag,
                        )
                        await conn.execute(
                            "INSERT INTO post_tags(post_id,tag_id) VALUES($1,$2)"
                            " ON CONFLICT DO NOTHING",
                            pid, tid,
                        )
    else:
        MEM_POSTS.insert(0, {
            "id": pid, "author_id": user["id"], "author": user["name"],
            "role": user["role"], "title": req.title, "body": req.body,
            "tags": [t.strip().lower() for t in req.tags if t.strip()][:10],
            "likes_count": 0, "status": status,
            "created_at": datetime.utcnow().isoformat(),
        })

    return {
        "id": pid, "status": status,
        "moderation": {
            "label": mod.label, "confidence": mod.confidence,
            "level": mod.level, "reasons": mod.reasons,
        },
    }

@app.put("/posts/{post_id}")
async def update_post(post_id: str, req: PostUpdateReq, user=Depends(get_user)):
    """Редактирование поста — только автор или админ."""
    if app.state.use_db and db.pool:
        post = await db.pool.fetchrow("SELECT author_id FROM posts WHERE id=$1", post_id)
        if not post:
            raise HTTPException(404, "Пост не найден")
        if post["author_id"] != user["id"] and user["role"] != "admin":
            raise HTTPException(403, "Можно редактировать только свои посты")
        fields, vals = [], []
        if req.title is not None:
            fields.append(f"title=${len(vals)+1}"); vals.append(req.title)
        if req.body is not None:
            fields.append(f"body=${len(vals)+1}"); vals.append(req.body)
        if req.status is not None and user["role"] == "admin":
            fields.append(f"status=${len(vals)+1}"); vals.append(req.status)
        if not fields:
            return {"updated": False}
        vals.append(post_id)
        await db.pool.execute(
            f"UPDATE posts SET {', '.join(fields)} WHERE id=${len(vals)}", *vals
        )
        return {"updated": True}

    # in-memory
    for p in MEM_POSTS:
        if p["id"] == post_id:
            if p["author_id"] != user["id"] and user["role"] != "admin":
                raise HTTPException(403, "Можно редактировать только свои посты")
            if req.title is not None: p["title"] = req.title
            if req.body  is not None: p["body"]  = req.body
            if req.tags  is not None:
                p["tags"] = [t.strip().lower() for t in req.tags if t.strip()][:10]
            if req.status is not None and user["role"] == "admin":
                p["status"] = req.status
            return {"updated": True}
    raise HTTPException(404, "Пост не найден")

@app.delete("/posts/{post_id}")
async def delete_post(post_id: str, user=Depends(get_user)):
    """Удалить пост — автор, модератор или админ."""
    if app.state.use_db and db.pool:
        post = await db.pool.fetchrow("SELECT author_id FROM posts WHERE id=$1", post_id)
        if not post:
            raise HTTPException(404, "Пост не найден")
        if post["author_id"] != user["id"] and user["role"] not in ("admin", "moderator"):
            raise HTTPException(403, "Нет прав на удаление")
        await db.pool.execute("DELETE FROM posts WHERE id=$1", post_id)
        return {"deleted": True}

    # in-memory
    for i, p in enumerate(MEM_POSTS):
        if p["id"] == post_id:
            if p["author_id"] != user["id"] and user["role"] not in ("admin", "moderator"):
                raise HTTPException(403, "Нет прав на удаление")
            MEM_POSTS.pop(i)
            # удалить связанные лайки
            for likes_set in MEM_LIKES.values():
                likes_set.discard(post_id)
            return {"deleted": True}
    raise HTTPException(404, "Пост не найден")

@app.post("/posts/{post_id}/like")
async def like_post(post_id: str, user=Depends(get_user)):
    """Toggle лайк (поставил/убрал)."""
    if app.state.use_db and db.pool:
        exists = await db.pool.fetchrow(
            "SELECT 1 FROM likes WHERE user_id=$1 AND post_id=$2",
            user["id"], post_id,
        )
        if exists:
            await db.pool.execute(
                "DELETE FROM likes WHERE user_id=$1 AND post_id=$2",
                user["id"], post_id)
            await db.pool.execute(
                "UPDATE posts SET likes_count=likes_count-1 WHERE id=$1", post_id)
            return {"liked": False}
        else:
            await db.pool.execute(
                "INSERT INTO likes(user_id,post_id) VALUES($1,$2) ON CONFLICT DO NOTHING",
                user["id"], post_id)
            await db.pool.execute(
                "UPDATE posts SET likes_count=likes_count+1 WHERE id=$1", post_id)
            return {"liked": True}

    # in-memory
    target = next((p for p in MEM_POSTS if p["id"] == post_id), None)
    if not target:
        raise HTTPException(404, "Пост не найден")
    user_likes = MEM_LIKES[user["id"]]
    if post_id in user_likes:
        user_likes.discard(post_id)
        target["likes_count"] = max(0, target["likes_count"] - 1)
        return {"liked": False, "likes_count": target["likes_count"]}
    else:
        user_likes.add(post_id)
        target["likes_count"] = target["likes_count"] + 1
        return {"liked": True, "likes_count": target["likes_count"]}

# ── Модерация ─────────────────────────────────────────────────────────────────
@app.post("/moderation/check")
async def moderation_check(req: ModReq, _=Depends(require_moderator)):
    """
    Классификация текста с подробным разбором решения.
    Возвращает вердикт, метрики, факторы риска и качество модели.
    """
    detail = app.state.mod.moderate_verbose(req.text)
    # Метрики качества модели на тестовой выборке
    try:
        detail["model_quality"] = app.state.mod.evaluate()
    except Exception:
        detail["model_quality"] = None
    return detail

# ── Рекомендации (РЕАЛЬНЫЕ, на основе тегов и лайков) ─────────────────────────

# ── Очередь модератора: посты на проверку ─────────────────────────────────────
@app.get("/posts/pending")
async def get_pending_posts(_=Depends(require_moderator)):
    """Возвращает посты со статусом 'suspicious' для проверки модератором."""
    if app.state.use_db and db.pool:
        rows = await db.pool.fetch("""
            SELECT p.id, p.author_id, u.name AS author, u.role,
                   p.title, p.body, p.status, p.mod_level, p.mod_conf,
                   p.created_at
            FROM posts p JOIN users u ON u.id = p.author_id
            WHERE p.status = 'suspicious'
            ORDER BY p.created_at DESC
        """)
        return {"posts": [dict(r) for r in rows]}
    # in-memory
    pending = [p for p in MEM_POSTS if p.get("status") == "suspicious"]
    return {"posts": pending}


@app.post("/posts/{post_id}/review")
async def review_post(post_id: str, req: PostReviewReq,
                      user=Depends(require_moderator)):
    """
    Модератор проверяет подозрительный пост.

    action:
      approve  — одобрить (status → approved, пост виден в ленте)
      reject   — отклонить (status → blocked, пост скрыт)
      recheck  — повторная модерация (перезапуск ML)

    Опционально: title/body — модератор может отредактировать перед одобрением.
    """
    if app.state.use_db and db.pool:
        post = await db.pool.fetchrow("SELECT * FROM posts WHERE id=$1", post_id)
        if not post:
            raise HTTPException(404, "Пост не найден")

        if req.action == "approve":
            new_title = req.title or post["title"]
            new_body = req.body or post["body"]
            await db.pool.execute(
                "UPDATE posts SET status='approved', title=$1, body=$2 WHERE id=$3",
                new_title, new_body, post_id)
            return {"id": post_id, "status": "approved", "action": "approve"}

        elif req.action == "reject":
            await db.pool.execute(
                "UPDATE posts SET status='blocked' WHERE id=$1", post_id)
            return {"id": post_id, "status": "blocked", "action": "reject"}

        elif req.action == "recheck":
            text = f"{req.title or post['title']}\n{req.body or post['body']}"
            mod = app.state.mod.moderate(text)
            new_status = "approved" if mod.label == "approved" else (
                "blocked" if mod.label == "blocked" else "suspicious")
            updates = ["status=$1", "mod_level=$2", "mod_conf=$3"]
            vals = [new_status, mod.level, float(mod.confidence)]
            if req.title:
                updates.append(f"title=${len(vals)+1}"); vals.append(req.title)
            if req.body:
                updates.append(f"body=${len(vals)+1}"); vals.append(req.body)
            vals.append(post_id)
            await db.pool.execute(
                f"UPDATE posts SET {','.join(updates)} WHERE id=${len(vals)}", *vals)
            return {"id": post_id, "status": new_status, "action": "recheck",
                    "moderation": {"label": mod.label, "confidence": mod.confidence}}

    else:
        # in-memory
        post = next((p for p in MEM_POSTS if p["id"] == post_id), None)
        if not post:
            raise HTTPException(404, "Пост не найден")

        if req.action == "approve":
            post["status"] = "approved"
            if req.title: post["title"] = req.title
            if req.body: post["body"] = req.body
            return {"id": post_id, "status": "approved", "action": "approve"}

        elif req.action == "reject":
            post["status"] = "blocked"
            return {"id": post_id, "status": "blocked", "action": "reject"}

        elif req.action == "recheck":
            text = f"{req.title or post['title']}\n{req.body or post['body']}"
            mod = app.state.mod.moderate(text)
            post["status"] = "approved" if mod.label == "approved" else (
                "blocked" if mod.label == "blocked" else "suspicious")
            if req.title: post["title"] = req.title
            if req.body: post["body"] = req.body
            return {"id": post_id, "status": post["status"], "action": "recheck",
                    "moderation": {"label": mod.label, "confidence": mod.confidence}}

    raise HTTPException(400, "Неизвестное действие")


@app.get("/recommendations/{user_id}")
async def recommendations(user_id: str, top_k: int = 5, user=Depends(get_user_optional)):
    """
    Персональные рекомендации на основе:
    1) Тегов постов которые написал сам пользователь
    2) Тегов постов которые он лайкнул
    Возвращаем посты других авторов с пересекающимися тегами.
    """
    if app.state.use_db and db.pool:
        # Соберём интересы — теги из своих постов и лайкнутых постов
        interest_tags = await db.pool.fetch("""
            SELECT DISTINCT t.name
            FROM tags t
            JOIN post_tags pt ON pt.tag_id = t.id
            JOIN posts p ON p.id = pt.post_id
            WHERE p.author_id = $1
               OR p.id IN (SELECT post_id FROM likes WHERE user_id = $1)
        """, user_id)
        tag_set = [r["name"] for r in interest_tags]
        if not tag_set:
            # cold start — самые популярные посты других авторов
            rows = await db.pool.fetch("""
                SELECT p.id, p.title, u.name AS author, p.likes_count,
                       array_agg(t.name) FILTER (WHERE t.name IS NOT NULL) AS tags
                FROM posts p
                JOIN users u ON u.id = p.author_id
                LEFT JOIN post_tags pt ON pt.post_id = p.id
                LEFT JOIN tags t ON t.id = pt.tag_id
                WHERE p.status='approved' AND p.author_id != $1
                GROUP BY p.id, u.id
                ORDER BY p.likes_count DESC
                LIMIT $2
            """, user_id, top_k)
            return {
                "user_id": user_id,
                "based_on": [],
                "recommendations": [
                    {"id": r["id"], "title": r["title"], "author": r["author"],
                     "likes_count": r["likes_count"], "tags": r["tags"] or [],
                     "score": 0.0, "reason": "Популярное"} for r in rows
                ],
            }

        # Подбор по пересечению тегов
        rows = await db.pool.fetch("""
            SELECT p.id, p.title, u.name AS author, p.likes_count,
                   array_agg(t.name) FILTER (WHERE t.name IS NOT NULL) AS tags,
                   COUNT(DISTINCT t.name) FILTER (WHERE t.name = ANY($2::text[])) AS overlap
            FROM posts p
            JOIN users u ON u.id = p.author_id
            LEFT JOIN post_tags pt ON pt.post_id = p.id
            LEFT JOIN tags t ON t.id = pt.tag_id
            WHERE p.status='approved'
              AND p.author_id != $1
              AND p.id NOT IN (SELECT post_id FROM likes WHERE user_id = $1)
            GROUP BY p.id, u.id
            HAVING COUNT(DISTINCT t.name) FILTER (WHERE t.name = ANY($2::text[])) > 0
            ORDER BY overlap DESC, p.likes_count DESC
            LIMIT $3
        """, user_id, tag_set, top_k)
        return {
            "user_id": user_id,
            "based_on": tag_set,
            "recommendations": [
                {"id": r["id"], "title": r["title"], "author": r["author"],
                 "likes_count": r["likes_count"], "tags": r["tags"] or [],
                 "score": float(r["overlap"]) / max(1, len(tag_set)),
                 "reason": f"Совпадают теги: {r['overlap']}/{len(tag_set)}"} for r in rows
            ],
        }

    # in-memory режим
    own_posts = [p for p in MEM_POSTS if p.get("author_id") == user_id]
    liked_post_ids = MEM_LIKES.get(user_id, set())
    liked_posts = [p for p in MEM_POSTS if p["id"] in liked_post_ids]

    # Собираем интересы из собственных постов + лайкнутых
    interest_counter = Counter()
    for p in own_posts + liked_posts:
        for tag in p.get("tags", []):
            interest_counter[tag] += 1

    if not interest_counter:
        # Cold start: топ постов других авторов по лайкам
        candidates = sorted(
            [p for p in MEM_POSTS
             if p.get("status") == "approved" and p.get("author_id") != user_id],
            key=lambda p: -p.get("likes_count", 0),
        )[:top_k]
        return {
            "user_id": user_id, "based_on": [],
            "recommendations": [
                {"id": p["id"], "title": p["title"], "author": p["author"],
                 "likes_count": p["likes_count"], "tags": p.get("tags", []),
                 "score": 0.0, "reason": "Популярное у других"}
                for p in candidates
            ],
        }

    # Скоринг по пересечению тегов
    interest_tags = list(interest_counter.keys())
    scored = []
    for p in MEM_POSTS:
        if p.get("status") != "approved": continue
        if p.get("author_id") == user_id: continue
        if p["id"] in liked_post_ids: continue
        ptags = set(p.get("tags", []))
        overlap = ptags.intersection(interest_tags)
        if overlap:
            score = sum(interest_counter[t] for t in overlap) / sum(interest_counter.values())
            scored.append((score, len(overlap), p))

    scored.sort(key=lambda x: (-x[0], -x[1], -x[2].get("likes_count", 0)))
    top = scored[:top_k]

    # Fallback: если по тегам ничего не нашлось — показываем популярные посты других авторов
    if not top:
        candidates = sorted(
            [p for p in MEM_POSTS
             if p.get("status") == "approved"
             and p.get("author_id") != user_id
             and p["id"] not in liked_post_ids],
            key=lambda p: -p.get("likes_count", 0),
        )[:top_k]
        return {
            "user_id": user_id, "based_on": interest_tags,
            "recommendations": [
                {"id": p["id"], "title": p["title"], "author": p["author"],
                 "likes_count": p["likes_count"], "tags": p.get("tags", []),
                 "score": 0.0, "reason": "Другие популярные публикации"}
                for p in candidates
            ],
        }

    return {
        "user_id": user_id,
        "based_on": interest_tags,
        "recommendations": [
            {"id": p["id"], "title": p["title"], "author": p["author"],
             "likes_count": p["likes_count"], "tags": p.get("tags", []),
             "score": round(score, 3),
             "reason": f"Совпадают теги: {n_match}/{len(interest_tags)}"}
            for score, n_match, p in top
        ],
    }

# ── Пользователи ──────────────────────────────────────────────────────────────
@app.get("/users")
async def get_users(_=Depends(require_admin)):
    if app.state.use_db and db.pool:
        # Реальные счётчики: посты и полученные лайки считаются из таблицы posts
        rows = await db.pool.fetch("""
            SELECT u.id, u.name, u.role, u.is_anomalous,
                   COUNT(p.id)                          AS posts_count,
                   COALESCE(SUM(p.likes_count), 0)::int AS likes_count
            FROM users u
            LEFT JOIN posts p ON p.author_id = u.id
            GROUP BY u.id, u.name, u.role, u.is_anomalous
            ORDER BY posts_count DESC
        """)
        return {"users": [dict(r) for r in rows]}
    # in-memory: пересчитываем посты/лайки на лету
    users_out = []
    for u in MEM_USERS.values():
        actual_posts = sum(1 for p in MEM_POSTS if p.get("author_id") == u["id"])
        likes_received = sum(p.get("likes_count", 0) for p in MEM_POSTS if p.get("author_id") == u["id"])
        users_out.append({
            "id": u["id"], "name": u["name"], "role": u["role"],
            "posts_count": actual_posts, "likes_count": likes_received,
            "is_anomalous": bool(u.get("is_anomalous", False)),
        })
    users_out.sort(key=lambda x: x["posts_count"], reverse=True)
    return {"users": users_out}

@app.post("/users", status_code=201)
async def create_user(req: UserCreateReq, _=Depends(require_admin)):
    """Только админ."""
    if app.state.use_db and db.pool:
        exists = await db.pool.fetchrow("SELECT 1 FROM users WHERE id=$1", req.username)
        if exists:
            raise HTTPException(400, "Пользователь с таким username уже существует")
        await db.pool.execute(
            "INSERT INTO users(id,name,role,password_hash) VALUES($1,$2,$3,$4)",
            req.username, req.name, req.role, _h(req.password),
        )
        return {"id": req.username, "name": req.name, "role": req.role}

    if req.username in MEM_USERS:
        raise HTTPException(400, "Пользователь с таким username уже существует")
    MEM_USERS[req.username] = {
        "id": req.username, "name": req.name, "role": req.role,
        "ph": _h(req.password), "posts": 0, "likes": 0,
    }
    return {"id": req.username, "name": req.name, "role": req.role}

@app.put("/users/{user_id}")
async def update_user(user_id: str, req: UserUpdateReq, _=Depends(require_admin)):
    """Только админ."""
    if app.state.use_db and db.pool:
        fields, vals = [], []
        if req.name is not None:
            fields.append(f"name=${len(vals)+1}"); vals.append(req.name)
        if req.role is not None:
            fields.append(f"role=${len(vals)+1}"); vals.append(req.role)
        if req.password is not None:
            fields.append(f"password_hash=${len(vals)+1}"); vals.append(_h(req.password))
        if not fields:
            return {"updated": False}
        vals.append(user_id)
        result = await db.pool.execute(
            f"UPDATE users SET {', '.join(fields)} WHERE id=${len(vals)}", *vals)
        return {"updated": result.endswith("1")}

    if user_id not in MEM_USERS:
        raise HTTPException(404, "Пользователь не найден")
    u = MEM_USERS[user_id]
    if req.name is not None: u["name"] = req.name
    if req.role is not None: u["role"] = req.role
    if req.password is not None: u["ph"] = _h(req.password)
    return {"updated": True}

@app.delete("/users/{user_id}")
async def delete_user(user_id: str, admin=Depends(require_admin)):
    """Только админ. Запрещено удалять самого себя."""
    if user_id == admin["id"]:
        raise HTTPException(400, "Нельзя удалить самого себя")
    if app.state.use_db and db.pool:
        result = await db.pool.execute("DELETE FROM users WHERE id=$1", user_id)
        return {"deleted": result.endswith("1")}

    if user_id not in MEM_USERS:
        raise HTTPException(404, "Пользователь не найден")
    del MEM_USERS[user_id]
    # Удалить связанные посты и лайки
    MEM_POSTS[:] = [p for p in MEM_POSTS if p.get("author_id") != user_id]
    MEM_LIKES.pop(user_id, None)
    for likes_set in MEM_LIKES.values():
        # лайки от удалённого юзера могли остаться у других — в данном случае не актуально
        pass
    return {"deleted": True}

# ── Обнаружение аномальных аккаунтов (SVD) ────────────────────────────────────
async def _gather_user_activities() -> List[UserActivity]:
    """
    Собирает данные всех аккаунтов системы для вычисления поведенческих
    признаков. Работает и с PostgreSQL, и с in-memory хранилищем.
    """
    activities: List[UserActivity] = []

    if app.state.use_db and db.pool:
        users_rows = await db.pool.fetch(
            "SELECT id, name, role, created_at FROM users ORDER BY id")
        for u in users_rows:
            posts_rows = await db.pool.fetch(
                "SELECT title, body, created_at, likes_count "
                "FROM posts WHERE author_id=$1", u["id"])
            likes_given_row = await db.pool.fetchrow(
                "SELECT COUNT(*) AS c FROM likes WHERE user_id=$1", u["id"])
            likes_given = int(likes_given_row["c"]) if likes_given_row else 0
            activities.append(UserActivity(
                user_id=u["id"], name=u["name"], role=u["role"],
                created_at=u["created_at"],
                posts=[dict(p) for p in posts_rows],
                likes_given=likes_given,
            ))
    else:
        for uid, u in MEM_USERS.items():
            user_posts = [p for p in MEM_POSTS if p.get("author_id") == uid]
            likes_given = len(MEM_LIKES.get(uid, set())) if isinstance(MEM_LIKES.get(uid), set) else 0
            activities.append(UserActivity(
                user_id=uid, name=u["name"], role=u["role"],
                posts=user_posts, likes_given=likes_given,
            ))
    return activities


@app.post("/users/detect-anomalies")
async def detect_anomalies(req: AnomalyDetectReq, _=Depends(require_admin)):
    """
    Запускает SVD-детектор аномальных аккаунтов.

    Реализует метод из статьи «Метод обнаружения аномальных аккаунтов
    в медицинских социальных платформах на основе сингулярного разложения
    матрицы поведения» (Аль-Раве М.И.Т., Макаров А.В.).

    Параметры:
        synthetic — добавить к реальным аккаунтам синтетические для
                    демонстрации работы алгоритма (если в системе мало данных).
        n_synthetic_legit — число «нормальных» синтетических аккаунтов.
        n_synthetic_anomalies — число «аномальных» синтетических аккаунтов.
        update_db — обновить поле is_anomalous в БД по результатам анализа
                    (только для реальных аккаунтов, синтетика игнорируется).
    """
    activities = await _gather_user_activities()

    if req.synthetic:
        synth = generate_synthetic_activities(
            n_legit=req.n_synthetic_legit,
            n_anomalies=req.n_synthetic_anomalies,
        )
        activities = activities + synth

    detector = AccountAnomalyDetector()
    result = detector.run(activities)

    # Обновление флага is_anomalous в БД (только для реальных аккаунтов)
    if req.update_db:
        real_results = [u for u in result["users"] if not u["is_synthetic"]]
        if app.state.use_db and db.pool:
            for u in real_results:
                await db.pool.execute(
                    "UPDATE users SET is_anomalous=$1 WHERE id=$2",
                    bool(u["is_anomalous"]), u["id"],
                )
        else:
            for u in real_results:
                if u["id"] in MEM_USERS:
                    MEM_USERS[u["id"]]["is_anomalous"] = bool(u["is_anomalous"])

    return result


@app.get("/users/{user_id}/behavioral-features")
async def get_user_features(user_id: str, _=Depends(require_admin)):
    """
    Возвращает 9 поведенческих признаков одного аккаунта без запуска SVD.
    Полезно для углублённого анализа конкретного пользователя.
    """
    activities = await _gather_user_activities()
    activity = next((a for a in activities if a.user_id == user_id), None)
    if activity is None:
        raise HTTPException(404, "Пользователь не найден")
    return {
        "id":          activity.user_id,
        "name":        activity.name,
        "role":        activity.role,
        "n_posts":     len(activity.posts),
        "likes_given": activity.likes_given,
        "features":    compute_features(activity),
    }


# ── Экспорт результатов в Excel (.xlsx) ───────────────────────────────────────
def _build_xlsx(sheets: List[dict], title: Optional[str] = None) -> BytesIO:
    """
    Собирает книгу Excel из списка листов. Каждый лист:
        {"name": str, "headers": [str], "rows": [[значения]]}
    Формат: жирная белая шапка на тёмно-синей заливке, тонкие рамки,
    автоширина столбцов, закреплённая строка заголовков.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="2E5090")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    title_font  = Font(bold=True, color="1B2A4A", size=13, name="Calibri")
    data_font   = Font(size=11, name="Calibri")
    thin        = Side(style="thin", color="BFBFBF")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for sh in sheets:
        name    = (sh.get("name") or "Лист")[:31]
        headers = sh.get("headers", [])
        rows    = sh.get("rows", [])
        ws = wb.create_sheet(title=name)
        start_row = 1

        # Необязательный заголовок листа над таблицей
        if title:
            ncols = max(len(headers), 1)
            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1, end_column=ncols)
            tc = ws.cell(row=1, column=1, value=title)
            tc.font = title_font
            tc.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 22
            start_row = 3   # пустая строка-отступ

        header_row_idx = start_row
        # Шапка
        if headers:
            for j, h in enumerate(headers, start=1):
                c = ws.cell(row=header_row_idx, column=j, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = center
                c.border = border
            ws.row_dimensions[header_row_idx].height = 30
            ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

        # Данные
        data_start = header_row_idx + (1 if headers else 0)
        for i, row in enumerate(rows):
            for j, val in enumerate(row, start=1):
                c = ws.cell(row=data_start + i, column=j, value=val)
                c.font = data_font
                c.border = border
                # Числа — по правому краю, текст — по левому
                if isinstance(val, (int, float)):
                    c.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center",
                                            wrap_text=True)

        # Автоширина столбцов
        ncols = max([len(headers)] + [len(r) for r in rows] + [1])
        for col in range(1, ncols + 1):
            letter = get_column_letter(col)
            max_len = 0
            for r in range(header_row_idx, data_start + len(rows)):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[letter].width = min(max(max_len + 3, 10), 48)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


@app.post("/export/xlsx")
async def export_xlsx(req: XlsxExportReq):
    """
    Формирует и отдаёт книгу Excel (.xlsx) из переданных данных.

    Тело запроса:
        filename — имя файла без расширения;
        title    — необязательный заголовок над таблицей на каждом листе;
        sheets   — список листов: {name, headers, rows}.

    Возвращает .xlsx с правильной структурой (ячейки, столбцы, строки),
    что исключает проблему распознавания разделителей при открытии CSV.
    """
    if not req.sheets:
        raise HTTPException(400, "Не переданы данные для экспорта")
    bio = _build_xlsx([s.model_dump() for s in req.sheets], req.title)
    safe_name = "".join(c for c in req.filename if c.isalnum() or c in "._-")[:80] or "export"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
    )


# ── Статистика ────────────────────────────────────────────────────────────────
@app.get("/stats")
async def stats(_=Depends(require_admin)):
    if app.state.use_db and db.pool:
        row = await db.pool.fetchrow("""
            SELECT
                (SELECT COUNT(*) FROM posts WHERE status='approved') AS posts,
                (SELECT COUNT(*) FROM users)                          AS users,
                (SELECT COUNT(*) FROM likes)                          AS likes,
                (SELECT ROUND(COUNT(*) FILTER (WHERE status='approved')::numeric
                              / NULLIF(COUNT(*),0)*100,1) FROM posts) AS moderation_pct
        """)
        return dict(row)
    total_likes = sum(len(s) for s in MEM_LIKES.values())
    return {
        "posts": len([p for p in MEM_POSTS if p.get("status") == "approved"]),
        "users": len(MEM_USERS), "likes": total_likes, "moderation_pct": 100.0,
    }

# ── Архитектуры (только админ) ────────────────────────────────────────────────
@app.get("/arch/list")
async def arch_list():
    return {"architectures": list_architectures()}

@app.get("/arch/info")
async def arch_info():
    from services.architectures import (
        MonolithApp, ThreeTierApp, MicroservicesApp, ServerlessApp
    )
    return {
        "monolith":      {"thread_pool": MonolithApp.THREAD_POOL,
                          "bottleneck": "общий lock БД"},
        "three_tier":    {"frontend_pool": ThreeTierApp.FRONTEND_POOL,
                          "backend_pool":  ThreeTierApp.BACKEND_POOL,
                          "db_pool":       ThreeTierApp.DB_POOL},
        "microservices": {"gateway_pool": MicroservicesApp.GATEWAY_BASE},
        "serverless":    {"cold_start_ms": ServerlessApp.COLD_START_LATENCY*1000},
    }

@app.get("/arch/{name}/loadtest")
async def arch_loadtest(name: str, n: int = 50, _=Depends(require_admin)):
    """Реальный concurrent load-test. Только админ."""
    arch = get_architecture(name)
    if not arch:
        raise HTTPException(404, f"Архитектура '{name}' не найдена")
    n = max(1, min(n, 2000))

    async def one_request():
        t0 = time.perf_counter()
        try:
            await asyncio.wait_for(arch.handle_request(), timeout=30.0)
            return (time.perf_counter() - t0), None
        except asyncio.TimeoutError:
            return 30.0, "timeout"
        except Exception as e:
            return time.perf_counter() - t0, str(e)[:80]

    t_start = time.perf_counter()
    results = await asyncio.gather(*[one_request() for _ in range(n)])
    total = time.perf_counter() - t_start

    times_ok = sorted([r[0] * 1000 for r in results if r[1] is None])
    errors = [r[1] for r in results if r[1] is not None]

    if not times_ok:
        return {"arch": name, "n": n, "rps": 0, "rt_avg": 0, "rt_p50": 0,
                "rt_p95": 0, "rt_p99": 0, "rt_min": 0, "rt_max": 0,
                "error_rate": 100.0, "duration_ms": round(total * 1000, 1)}

    def pct(p):
        i = int(len(times_ok) * p)
        return times_ok[min(i, len(times_ok)-1)]

    return {
        "arch": name, "n": n,
        "rps": round(n / total, 1) if total else 0,
        "rt_avg": round(sum(times_ok) / len(times_ok), 1),
        "rt_min": round(times_ok[0], 1),
        "rt_p50": round(pct(0.50), 1),
        "rt_p95": round(pct(0.95), 1),
        "rt_p99": round(pct(0.99), 1),
        "rt_max": round(times_ok[-1], 1),
        "error_rate": round(len(errors) / n * 100, 2),
        "duration_ms": round(total * 1000, 1),
    }
